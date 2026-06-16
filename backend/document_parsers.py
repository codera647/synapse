"""
backend/document_parsers.py

Multi-format front-ends that normalize non-PDF documents into the SAME text IR the rest of
the pipeline already consumes (chunk -> embed -> retrieve), so adding a format is just a new
parser — nothing downstream changes.

IR returned by parse_to_ir():
  {
    "format": "csv" | "xlsx" | "code" | "text",
    "pages":  [ {"page": int, "blocks": [block, ...]} ],
    "links":  [],
  }
where block = {
    "block_id": str, "kind": "text", "text": str,
    "bbox_img": [x1,y1,x2,y2],   # synthetic, increasing y => preserves reading order for chunker
    "locator": str,             # citation locator: "Sheet!rows a-b" / "file.py:10-40" / filename
    "force_chunk": bool,        # code: one chunk per function/class
    "section_heading": str|None,
}
Returns None if the format is unsupported or parsing fails (caller skips the doc).
"""

from __future__ import annotations

import csv
import io
import os
import re
from typing import Any, Dict, List, Optional

_CODE_EXTS = {
    ".py", ".js", ".ts", ".tsx", ".jsx", ".java", ".go", ".rs", ".rb", ".php", ".c", ".cc",
    ".cpp", ".h", ".hpp", ".cs", ".swift", ".kt", ".scala", ".sh", ".bash", ".sql", ".r",
    ".m", ".lua", ".pl", ".dart", ".vue", ".css", ".scss",
}


def _block(idx: int, text: str, locator: str = "", force_chunk: bool = False,
           section_heading: Optional[str] = None) -> Dict[str, Any]:
    return {
        "block_id": f"b{idx:04d}",
        "kind": "text",  # chunk_worker only ingests text blocks; tables are serialized as text
        "text": text,
        "bbox_img": [0, idx * 10, 1000, idx * 10 + 10],
        "locator": locator or "",
        "force_chunk": bool(force_chunk),
        "section_heading": section_heading,
    }


def parse_to_ir(mime_type: Optional[str], filename: Optional[str], file_bytes: bytes) -> Optional[Dict[str, Any]]:
    name = (filename or "").lower()
    mt = (mime_type or "").lower()
    try:
        if name.endswith(".csv") or "csv" in mt:
            return _parse_csv(file_bytes, filename or "data.csv")
        if name.endswith((".xlsx", ".xlsm")) or "spreadsheetml" in mt or "ms-excel" in mt:
            return _parse_xlsx(file_bytes, filename or "workbook.xlsx")
        if name.endswith(".docx") or "wordprocessingml" in mt:
            return _parse_docx(file_bytes, filename or "document.docx")
        if any(name.endswith(e) for e in _CODE_EXTS):
            return _parse_code(file_bytes, filename or "file.txt")
        if name.endswith((".md", ".markdown")) or "markdown" in mt:
            return _parse_text(file_bytes, filename or "doc.md", markdown=True)
        if name.endswith((".txt", ".text", ".log")) or mt.startswith("text/plain"):
            return _parse_text(file_bytes, filename or "doc.txt")
    except Exception:
        return None
    return None


def _decode(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8", errors="replace")


def _rows_to_blocks(header: List[str], rows: List[List[str]], sheet: str, start_idx: int = 0) -> List[Dict[str, Any]]:
    group = int(os.getenv("PARSE_TABLE_ROWS_PER_CHUNK", "30"))
    hdr = " | ".join(h for h in header)
    blocks: List[Dict[str, Any]] = []
    bi = start_idx
    for start in range(0, len(rows), group):
        cr = rows[start : start + group]
        lines = [hdr] + [" | ".join(c for c in r) for r in cr]
        loc = f"{sheet}!rows {start + 2}-{start + 1 + len(cr)}"
        blocks.append(_block(bi, "\n".join(lines), locator=loc, force_chunk=True))
        bi += 1
    return blocks


def _parse_docx(file_bytes: bytes, filename: str) -> Optional[Dict[str, Any]]:
    """Word .docx -> text IR: paragraphs (headings kept as section headings) + tables serialized as
    text (same convention as xlsx). Needs python-docx; returns None if unavailable. .doc (old binary)
    is NOT supported — convert to .docx or PDF."""
    try:
        from docx import Document  # python-docx
    except Exception:
        return None
    try:
        doc = Document(io.BytesIO(file_bytes))
    except Exception:
        return None
    blocks: List[Dict[str, Any]] = []
    bi = 0
    for p in doc.paragraphs:
        txt = (p.text or "").strip()
        if not txt:
            continue
        style = (getattr(p.style, "name", "") or "").lower()
        heading = txt[:80] if "heading" in style or "title" in style else None
        blocks.append(_block(bi, txt, locator=filename, section_heading=heading))
        bi += 1
    for ti, table in enumerate(getattr(doc, "tables", []) or []):
        rows = [[(c.text or "").strip() for c in row.cells] for row in table.rows]
        rows = [r for r in rows if any(c for c in r)]
        if not rows:
            continue
        lines = [" | ".join(rows[0])] + [" | ".join(r) for r in rows[1:]]
        blocks.append(_block(bi, "\n".join(lines), locator=f"{filename} table {ti + 1}", force_chunk=True))
        bi += 1
    if not blocks:
        return None
    return {"format": "docx", "pages": [{"page": 0, "blocks": blocks}], "links": []}


def _parse_csv(file_bytes: bytes, filename: str) -> Optional[Dict[str, Any]]:
    rows = list(csv.reader(io.StringIO(_decode(file_bytes))))
    rows = [[("" if c is None else str(c)) for c in r] for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        return None
    header, data = rows[0], rows[1:]
    summary = _block(0, f"CSV {filename}: columns {', '.join(h for h in header if h)}; {len(data)} rows.", locator=filename)
    blocks = [summary] + _rows_to_blocks(header, data, sheet=filename, start_idx=1)
    return {"format": "csv", "pages": [{"page": 0, "blocks": blocks}], "links": []}


def _parse_xlsx(file_bytes: bytes, filename: str) -> Optional[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    pages: List[Dict[str, Any]] = []
    for pi, ws in enumerate(wb.worksheets):
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
        rows = [r for r in rows if any((c or "").strip() for c in r)]
        if not rows:
            continue
        header, data = rows[0], rows[1:]
        summary = _block(0, f"Sheet '{ws.title}': columns {', '.join(h for h in header if h)}; {len(data)} rows.", locator=ws.title)
        blocks = [summary] + _rows_to_blocks(header, data, sheet=ws.title, start_idx=1)
        pages.append({"page": pi, "blocks": blocks})
    if not pages:
        return None
    return {"format": "xlsx", "pages": pages, "links": []}


_DEF_RE = re.compile(
    r"^\s*(?:export\s+)?(?:default\s+)?(?:async\s+)?"
    r"(?:def|class|function|func|fn|interface|type|struct|enum|impl|module|"
    r"public|private|protected|static|const\s+\w+\s*=\s*(?:async\s*)?\()"
    r"\b"
)


def _parse_code(file_bytes: bytes, filename: str) -> Optional[Dict[str, Any]]:
    lines = _decode(file_bytes).splitlines()
    if not lines:
        return None
    bounds = sorted({i for i, l in enumerate(lines) if _DEF_RE.match(l)})
    blocks: List[Dict[str, Any]] = []
    bi = 0
    if len(bounds) >= 2:
        starts = sorted(set([0] + bounds + [len(lines)]))
        for j in range(len(starts) - 1):
            seg_lines = lines[starts[j] : starts[j + 1]]
            seg = "\n".join(seg_lines).strip()
            if not seg:
                continue
            sym = next((l.strip() for l in seg_lines if l.strip()), "")[:80]
            blocks.append(
                _block(bi, seg, locator=f"{filename}:{starts[j] + 1}-{starts[j + 1]}", force_chunk=True, section_heading=sym)
            )
            bi += 1
    else:
        win = int(os.getenv("PARSE_CODE_WINDOW_LINES", "60"))
        for start in range(0, len(lines), win):
            seg = "\n".join(lines[start : start + win]).strip()
            if not seg:
                continue
            end = min(len(lines), start + win)
            blocks.append(_block(bi, seg, locator=f"{filename}:{start + 1}-{end}", force_chunk=True))
            bi += 1
    if not blocks:
        return None
    return {"format": "code", "pages": [{"page": 0, "blocks": blocks}], "links": []}


def _parse_text(file_bytes: bytes, filename: str, markdown: bool = False) -> Optional[Dict[str, Any]]:
    txt = _decode(file_bytes)
    paras = [p.strip() for p in re.split(r"\n\s*\n", txt) if p.strip()]
    if not paras:
        return None
    blocks: List[Dict[str, Any]] = []
    for i, p in enumerate(paras):
        heading = p.lstrip("#").strip()[:80] if (markdown and p.startswith("#")) else None
        blocks.append(_block(i, p, locator=filename, section_heading=heading))
    return {"format": "text", "pages": [{"page": 0, "blocks": blocks}], "links": []}
