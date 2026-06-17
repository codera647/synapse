"""
backend/agent_data.py

Data acquisition for Agent mode.

- load_tabular(): EXACT re-parse of a structured file (xlsx/csv) into typed columns/series for
  charting. Unlike document_parsers._parse_xlsx (which flattens to text for retrieval), this keeps
  NATIVE cell types so numeric series are real numbers, not strings.
- R2 helpers: fetch_r2_bytes / put_r2_bytes / put_r2_json / put_r2_png (Agent mode writes artifacts
  under the `agents/` prefix and reads uploads under `agent-uploads/`).
- Source resolution: load a library doc's raw file (documents.storage_path_raw) or a runtime upload
  (agent_uploads.storage_key) and parse it.
"""

from __future__ import annotations

import datetime as _dt
import io
import json
import os
from decimal import Decimal as _Decimal
from typing import Any, Dict, List, Optional


def _json_safe(v: Any) -> Any:
    """Coerce a parsed cell value to a JSON-serializable type. Excel date/time cells
    come back as datetime objects (and numbers can be Decimal), which break the JSONB
    insert into agent_uploads.preview and any spec serialization."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, _Decimal):
        return float(v)
    if isinstance(v, bytes):
        return v.decode("utf-8", "replace")
    return str(v)

import boto3

from chat_runtime import supabase  # service-role client, already constructed

_R2_BUCKET = os.getenv("R2_BUCKET") or ""
_s3 = boto3.client(
    "s3",
    endpoint_url=(os.getenv("R2_ENDPOINT") or None),
    aws_access_key_id=(os.getenv("R2_ACCESS_KEY") or None),
    aws_secret_access_key=(os.getenv("R2_SECRET_KEY") or None),
)

_STRUCTURED_EXT = (".xlsx", ".xlsm", ".csv")


# ── R2 ────────────────────────────────────────────────────────────────────────────────────────
def fetch_r2_bytes(key: str) -> bytes:
    obj = _s3.get_object(Bucket=_R2_BUCKET, Key=key)
    return obj["Body"].read()


def put_r2_bytes(key: str, data: bytes, content_type: str = "application/octet-stream") -> None:
    _s3.put_object(Bucket=_R2_BUCKET, Key=key, Body=data, ContentType=content_type)


def put_r2_json(key: str, payload: Dict[str, Any]) -> None:
    put_r2_bytes(key, json.dumps(payload).encode("utf-8"), "application/json")


def put_r2_png(key: str, png_bytes: bytes) -> None:
    put_r2_bytes(key, png_bytes, "image/png")


# ── Tabular parsing ──────────────────────────────────────────────────────────────────────────
def is_structured(filename: str, mime_type: str = "") -> bool:
    name = (filename or "").lower()
    mt = (mime_type or "").lower()
    return name.endswith(_STRUCTURED_EXT) or "spreadsheet" in mt or "csv" in mt or "ms-excel" in mt


def _to_number(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").replace("%", "").strip())
    except Exception:
        return None


def _infer_dtype(values: List[Any]) -> str:
    nonnull = [v for v in values if v is not None and str(v).strip() != ""]
    if not nonnull:
        return "string"
    if all(_to_number(v) is not None for v in nonnull):
        return "number"
    # cheap date sniff (don't hard-depend on dateutil)
    looks_date = 0
    for v in nonnull[:20]:
        s = str(v)
        if any(ch in s for ch in ("-", "/", ":")) and any(c.isdigit() for c in s) and len(s) <= 24:
            looks_date += 1
    if looks_date >= max(1, min(len(nonnull), 20)) * 0.8:
        return "date"
    return "string"


def _finalize(columns: List[str], rows: List[List[Any]], sheet: str, sheets: List[str],
              max_full_rows: int) -> Dict[str, Any]:
    truncated = len(rows) > max_full_rows
    rows = rows[:max_full_rows]
    series: Dict[str, List[Any]] = {c: [] for c in columns}
    for r in rows:
        for i, c in enumerate(columns):
            series[c].append(r[i] if i < len(r) else None)
    dtypes = {c: _infer_dtype(series[c]) for c in columns}
    # cast numeric columns to real numbers for charting; make everything else
    # JSON-safe (datetime/date/Decimal cells would otherwise break the preview insert)
    for c in columns:
        if dtypes[c] == "number":
            series[c] = [_to_number(v) for v in series[c]]
        else:
            series[c] = [_json_safe(v) for v in series[c]]
    sample_rows = [{c: (series[c][i] if i < len(series[c]) else None) for c in columns}
                   for i in range(min(20, len(rows)))]
    return {
        "sheets": sheets,
        "active_sheet": sheet,
        "columns": columns,
        "dtypes": dtypes,
        "row_count": len(rows),
        "sample_rows": sample_rows,
        "series": series,
        "truncated": truncated,
    }


def _load_xlsx(raw: bytes, sheet: Optional[str], max_full_rows: int) -> Optional[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    if not sheets:
        return None
    ws = wb[sheet] if (sheet and sheet in sheets) else wb[sheets[0]]
    all_rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    all_rows = [r for r in all_rows if any((str(c).strip() if c is not None else "") for c in r)]
    if not all_rows:
        return None
    header = [("" if c is None else str(c)).strip() or f"col{i+1}" for i, c in enumerate(all_rows[0])]
    return _finalize(header, all_rows[1:], ws.title, sheets, max_full_rows)


def _load_csv(raw: bytes, max_full_rows: int) -> Optional[Dict[str, Any]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        return None
    import csv as _csv

    reader = list(_csv.reader(io.StringIO(text)))
    reader = [r for r in reader if any((c or "").strip() for c in r)]
    if not reader:
        return None
    header = [(c or "").strip() or f"col{i+1}" for i, c in enumerate(reader[0])]
    return _finalize(header, reader[1:], "Sheet1", ["Sheet1"], max_full_rows)


def load_tabular(raw_bytes: bytes, filename: str, mime_type: str = "",
                 sheet: Optional[str] = None, max_full_rows: int = 50000) -> Optional[Dict[str, Any]]:
    """Exact re-parse of a structured file -> typed columns/series. None if not structured/parseable."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in (mime_type or "").lower() or "ms-excel" in (mime_type or "").lower():
        return _load_xlsx(raw_bytes, sheet, max_full_rows)
    if name.endswith(".csv") or "csv" in (mime_type or "").lower():
        return _load_csv(raw_bytes, max_full_rows)
    return None


# ── Source resolution ───────────────────────────────────────────────────────────────────────
def fetch_upload_row(upload_id: str, organization_id: str) -> Optional[Dict[str, Any]]:
    res = (supabase.table("agent_uploads")
           .select("id, filename, mime_type, storage_key, kind, library_id")
           .eq("id", upload_id).eq("organization_id", organization_id).limit(1).execute())
    return (res.data or [None])[0]


def fetch_doc_row(doc_id: str, organization_id: str) -> Optional[Dict[str, Any]]:
    res = (supabase.table("documents")
           .select("id, title, storage_path_raw, mime_type, library_id")
           .eq("id", doc_id).eq("organization_id", organization_id).limit(1).execute())
    return (res.data or [None])[0]


def load_structured(organization_id: str, source_kind: str, source_ref: str,
                    sheet: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """Resolve a source ref to a parsed table. Returns the load_tabular dict + {source_name}, or None
    if the source isn't a structured file we can re-parse."""
    if source_kind == "upload":
        row = fetch_upload_row(source_ref, organization_id)
        if not row or not row.get("storage_key"):
            return None
        raw = fetch_r2_bytes(row["storage_key"])
        out = load_tabular(raw, row.get("filename") or "", row.get("mime_type") or "", sheet=sheet)
    elif source_kind in ("library_structured", "library"):
        row = fetch_doc_row(source_ref, organization_id)
        if not row or not row.get("storage_path_raw"):
            return None
        raw = fetch_r2_bytes(row["storage_path_raw"])
        out = load_tabular(raw, row.get("title") or "", row.get("mime_type") or "", sheet=sheet)
    else:
        return None
    if out:
        out["source_name"] = (row.get("filename") or row.get("title") or source_ref)
        out["source_ref"] = source_ref
        out["source_kind"] = source_kind
    return out


def list_library_documents(organization_id: str, library_ids: List[str], limit: int = 200) -> List[Dict[str, Any]]:
    """Enumerate docs across the selected libraries (for the planner's source summary)."""
    if not library_ids:
        return []
    res = (supabase.table("documents")
           .select("id, title, mime_type, storage_path_raw, library_id")
           .eq("organization_id", organization_id)
           .in_("library_id", library_ids).limit(limit).execute())
    return res.data or []
